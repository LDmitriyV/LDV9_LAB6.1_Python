import os
import sys
from pathlib import Path
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docxtpl import DocxTemplate
from openpyxl import load_workbook

class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('uis/main.ui', self)
        self.setWindowIcon(QtGui.QIcon('images/icon.png'))
        self.setWindowTitle('LDV9_LAB6.1 Работа с массивами и файлами в Python')

        self.btn_wrd.clicked.connect(self.execute_wrd)
        self.btn_xl.clicked.connect(self.execute_xl)

    def execute_wrd(self):
        document_path = Path(__file__).parent / "card.docx"
        doc = DocxTemplate(document_path)
        context = {"your_FIO": self.lineEdit.text(),
                   "job": self.lineEdit_2.text(),
                   "post": self.lineEdit_3.text(),
                   "adress": self.lineEdit_4.text(),
                   "city": self.lineEdit_5.text(),
                   "working_time": self.lineEdit_6.text(),
                   "phone_number": self.lineEdit_7.text(),
                   "website": self.lineEdit_8.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "generated_card.docx")
        os.system('start generated_card.docx')

    def execute_xl(self):
        fn = 'card.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['A1'] = self.lineEdit.text()
        ws['B3'] = self.lineEdit_2.text()
        ws['B4'] = self.lineEdit_3.text()
        ws['B5'] = self.lineEdit_4.text()
        ws['B6'] = self.lineEdit_5.text()
        ws['B7'] = self.lineEdit_6.text()
        ws['B8'] = self.lineEdit_7.text()
        ws['B9'] = self.lineEdit_8.text()

        wb.save(Path(__file__).parent / "generated_card.xlsx")
        wb.close()
        os.system('start generated_card.xlsx')

def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
